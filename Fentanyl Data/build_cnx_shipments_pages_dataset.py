#!/usr/bin/env python3
from __future__ import annotations

import re
from collections import defaultdict
from difflib import SequenceMatcher
from pathlib import Path

from openpyxl import load_workbook
import pandas as pd


US_STATE_ABBR_TO_NAME = {
    "AL": "Alabama",
    "AK": "Alaska",
    "AZ": "Arizona",
    "AR": "Arkansas",
    "CA": "California",
    "CO": "Colorado",
    "CT": "Connecticut",
    "DE": "Delaware",
    "FL": "Florida",
    "GA": "Georgia",
    "HI": "Hawaii",
    "ID": "Idaho",
    "IL": "Illinois",
    "IN": "Indiana",
    "IA": "Iowa",
    "KS": "Kansas",
    "KY": "Kentucky",
    "LA": "Louisiana",
    "ME": "Maine",
    "MD": "Maryland",
    "MA": "Massachusetts",
    "MI": "Michigan",
    "MN": "Minnesota",
    "MS": "Mississippi",
    "MO": "Missouri",
    "MT": "Montana",
    "NE": "Nebraska",
    "NV": "Nevada",
    "NH": "New Hampshire",
    "NJ": "New Jersey",
    "NM": "New Mexico",
    "NY": "New York",
    "NC": "North Carolina",
    "ND": "North Dakota",
    "OH": "Ohio",
    "OK": "Oklahoma",
    "OR": "Oregon",
    "PA": "Pennsylvania",
    "RI": "Rhode Island",
    "SC": "South Carolina",
    "SD": "South Dakota",
    "TN": "Tennessee",
    "TX": "Texas",
    "UT": "Utah",
    "VT": "Vermont",
    "VA": "Virginia",
    "WA": "Washington",
    "WV": "West Virginia",
    "WI": "Wisconsin",
    "WY": "Wyoming",
    "DC": "District of Columbia",
}

STATE_CODES = set(US_STATE_ABBR_TO_NAME.keys())
STATE_NAME_TO_ABBR = {name.upper(): abbr for abbr, name in US_STATE_ABBR_TO_NAME.items()}
STATE_NAME_TO_ABBR.update(
    {
        "WASHINGTON DC": "DC",
        "WASHINGTON D C": "DC",
        "WASHINGTON, DC": "DC",
        "DISTRICT OF COLUMBIA": "DC",
    }
)

STATE_ABBR_PATTERN = "|".join(sorted(STATE_CODES, key=len, reverse=True))
STATE_NAME_PATTERN = "|".join(re.escape(name) for name in sorted(STATE_NAME_TO_ABBR.keys(), key=len, reverse=True))
ZIP_STATE_RE = re.compile(rf"\b({STATE_ABBR_PATTERN})\s+\d{{5}}(?:-\d{{4}})?\b", flags=re.IGNORECASE)
COMMA_STATE_RE = re.compile(rf",\s*({STATE_ABBR_PATTERN})\s*(?:,|$)", flags=re.IGNORECASE)
STATE_NAME_RE = re.compile(rf"\b({STATE_NAME_PATTERN})\b", flags=re.IGNORECASE)
ALIAS_SPLIT_RE = re.compile(r"[;\n|]+")


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).replace("\xa0", " ").lower()
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def normalize_hs6(value: object) -> str | None:
    if pd.isna(value):
        return None
    text = str(value).strip()
    if text == "" or text.lower() in {"null", "none", "nan"}:
        return None
    digits_only = "".join(ch for ch in text if ch.isdigit())
    if len(digits_only) < 6:
        return None
    return digits_only[:6]


def extract_state_from_address(address: object) -> str | None:
    if not isinstance(address, str):
        return None
    text = address.strip()
    if not text:
        return None

    normalized = re.sub(r"\s+", " ", text.upper().replace(".", " ")).strip()

    zip_match = ZIP_STATE_RE.search(normalized)
    if zip_match:
        return zip_match.group(1).upper()

    comma_match = COMMA_STATE_RE.search(normalized)
    if comma_match:
        return comma_match.group(1).upper()

    name_match = STATE_NAME_RE.search(normalized)
    if name_match:
        return STATE_NAME_TO_ABBR[name_match.group(1).upper()]

    tokens = [segment.strip() for segment in normalized.split(",") if segment.strip()]
    for token in reversed(tokens):
        if token in STATE_CODES:
            return token
        if token in STATE_NAME_TO_ABBR:
            return STATE_NAME_TO_ABBR[token]

        token_abbr_match = re.match(r"^([A-Z]{2})(?:\s+\d{5}(?:-\d{4})?)?$", token)
        if token_abbr_match:
            candidate = token_abbr_match.group(1)
            if candidate in STATE_CODES:
                return candidate

        token_name_zip_match = re.match(r"^([A-Z ]+)\s+\d{5}(?:-\d{4})?\b", token)
        if token_name_zip_match:
            candidate = token_name_zip_match.group(1).strip()
            if candidate in STATE_NAME_TO_ABBR:
                return STATE_NAME_TO_ABBR[candidate]

    return None


def load_base_shipments(input_csv: Path) -> pd.DataFrame:
    usecols = [
        "transaction_id",
        "receiver_country_iso2",
        "receiver_address",
        "transaction_date",
        "goods_description",
        "hs_code",
        "hs6_code",
        "predicted_hs6_codes_top_1",
        "kg_net",
        "kg_gross",
        "analytical_value_usd",
        "est_usd_value",
    ]
    df = pd.read_csv(input_csv, usecols=usecols, low_memory=False)
    df["receiver_country_iso2"] = df["receiver_country_iso2"].astype(str).str.strip().str.upper()
    df = df[df["receiver_country_iso2"] == "US"].copy()

    df["transaction_date"] = pd.to_datetime(df["transaction_date"], errors="coerce")
    df["year"] = df["transaction_date"].dt.year.astype("Int64")

    hs6_final = pd.Series([None] * len(df), index=df.index, dtype="object")
    for col in ["hs6_code", "hs_code", "predicted_hs6_codes_top_1"]:
        hs6_candidate = df[col].apply(normalize_hs6)
        hs6_final = hs6_final.fillna(hs6_candidate)
    df["hs6"] = hs6_final

    df["receiver_address"] = df["receiver_address"].astype(str)
    df["state_abbr"] = df["receiver_address"].apply(extract_state_from_address)
    df["state_name"] = df["state_abbr"].map(US_STATE_ABBR_TO_NAME)

    df["kg_net"] = pd.to_numeric(df["kg_net"], errors="coerce")
    df["kg_gross"] = pd.to_numeric(df["kg_gross"], errors="coerce")
    df["quantity_kg"] = df["kg_net"].where(df["kg_net"].notna(), df["kg_gross"])

    df["analytical_value_usd"] = pd.to_numeric(df["analytical_value_usd"], errors="coerce")
    df["est_usd_value"] = pd.to_numeric(df["est_usd_value"], errors="coerce")
    df["value_usd"] = df["analytical_value_usd"].where(
        df["analytical_value_usd"].notna(), df["est_usd_value"]
    )

    return df


def build_pages_hs6_dataset(base_df: pd.DataFrame, output_csv: Path) -> None:
    base = base_df.dropna(subset=["year", "hs6", "state_abbr"]).copy()
    base["year"] = base["year"].astype(int)
    base["hs6"] = base["hs6"].astype(str)

    grouped = (
        base.groupby(["state_abbr", "state_name", "year", "hs6"], as_index=False)
        .agg(
            shipment_records=("transaction_id", "count"),
            total_quantity_kg=("quantity_kg", "sum"),
            total_value_usd=("value_usd", "sum"),
        )
        .fillna({"total_quantity_kg": 0.0, "total_value_usd": 0.0})
        .sort_values(["year", "state_abbr", "hs6"])
        .reset_index(drop=True)
    )

    output_csv.parent.mkdir(parents=True, exist_ok=True)
    grouped.to_csv(output_csv, index=False)

    print(f"Wrote: {output_csv}")
    print(f"Rows: {len(grouped):,}")
    print(f"Mapped source records used: {len(base):,}")
    print(f"Unique HS6: {grouped['hs6'].nunique():,}")
    print(f"Year range: {grouped['year'].min()}-{grouped['year'].max()}")


def load_precursor_alias_records(precursor_workbook: Path) -> list[dict[str, object]]:
    wb = load_workbook(precursor_workbook, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)

    headers = [str(v).strip() if v is not None else "" for v in next(rows)]
    idx = {name: i for i, name in enumerate(headers)}
    chem_idx = idx.get("chemical_name_list")
    alt_idx = idx.get("alt_names")
    if chem_idx is None:
        raise ValueError("Workbook is missing chemical_name_list column")

    registry: dict[str, dict[str, object]] = {}
    for row in rows:
        values = list(row) if row is not None else []
        chemical_raw = values[chem_idx] if chem_idx < len(values) else None
        chemical_name = str(chemical_raw).strip() if chemical_raw is not None else ""
        if not chemical_name:
            continue
        chemical_name = re.sub(r"\s+", " ", chemical_name.replace("\xa0", " ")).strip()
        chemical_norm = normalize_text(chemical_name)
        if not chemical_norm:
            continue

        rec = registry.setdefault(
            chemical_norm,
            {"chemical_name": chemical_name, "aliases": set()},
        )
        rec["aliases"].add(chemical_name)

        if alt_idx is not None and alt_idx < len(values):
            alt_raw = values[alt_idx]
            alt_text = str(alt_raw).strip() if alt_raw is not None else ""
            if alt_text:
                for alias in ALIAS_SPLIT_RE.split(alt_text):
                    alias = re.sub(r"\s+", " ", alias.replace("\xa0", " ")).strip()
                    if alias:
                        rec["aliases"].add(alias)

    alias_records: list[dict[str, object]] = []
    for rec in registry.values():
        chemical_name = rec["chemical_name"]
        aliases = rec["aliases"]
        for alias in aliases:
            alias_norm = normalize_text(alias)
            if len(alias_norm) < 3:
                continue
            tokens = alias_norm.split()
            alias_records.append(
                {
                    "chemical_name": chemical_name,
                    "alias": alias,
                    "alias_norm": alias_norm,
                    "tokens": tokens,
                    "token_len": len(tokens),
                }
            )

    alias_records.sort(
        key=lambda r: (
            r["chemical_name"],
            -len(r["alias_norm"]),
            r["alias_norm"],
        )
    )
    return alias_records


def build_alias_index(alias_records: list[dict[str, object]]) -> tuple[dict[str, set[int]], int]:
    token_to_alias_ids: dict[str, set[int]] = defaultdict(set)
    max_alias_token_len = 1
    for i, rec in enumerate(alias_records):
        tokens = rec["tokens"]
        max_alias_token_len = max(max_alias_token_len, len(tokens))
        for token in tokens:
            if len(token) >= 3:
                token_to_alias_ids[token].add(i)
    return token_to_alias_ids, max_alias_token_len


def match_chemical_from_goods_description(
    goods_description: object,
    alias_records: list[dict[str, object]],
    token_to_alias_ids: dict[str, set[int]],
    max_alias_token_len: int,
) -> dict[str, object] | None:
    desc_norm = normalize_text(goods_description)
    if not desc_norm:
        return None
    desc_tokens = desc_norm.split()
    if not desc_tokens:
        return None

    candidate_ids: set[int] = set()
    for token in set(desc_tokens):
        candidate_ids.update(token_to_alias_ids.get(token, set()))
    if not candidate_ids:
        return None

    padded_desc = f" {desc_norm} "

    best_exact: dict[str, object] | None = None
    best_exact_key: tuple[float, int] = (-1.0, -1)
    for idx in candidate_ids:
        rec = alias_records[idx]
        alias_norm = rec["alias_norm"]
        if f" {alias_norm} " in padded_desc:
            key = (1.0, len(alias_norm))
            if key > best_exact_key:
                best_exact_key = key
                best_exact = {
                    "chemical_name": rec["chemical_name"],
                    "matched_alias": rec["alias"],
                    "match_score": 1.0,
                    "match_type": "exact",
                }
    if best_exact is not None:
        return best_exact

    max_window_len = min(max_alias_token_len + 1, 8, len(desc_tokens))
    windows_by_len: dict[int, list[str]] = {}
    for n in range(1, max_window_len + 1):
        windows_by_len[n] = [" ".join(desc_tokens[i : i + n]) for i in range(len(desc_tokens) - n + 1)]

    best_fuzzy: dict[str, object] | None = None
    best_fuzzy_key: tuple[float, int] = (0.0, -1)
    for idx in candidate_ids:
        rec = alias_records[idx]
        alias_norm = rec["alias_norm"]
        alias_token_len = rec["token_len"]

        threshold = 0.88 if alias_token_len >= 2 else (0.92 if len(alias_norm) >= 6 else 0.97)
        local_best = 0.0
        for n in (alias_token_len - 1, alias_token_len, alias_token_len + 1):
            if n <= 0:
                continue
            for window in windows_by_len.get(n, []):
                score = SequenceMatcher(None, alias_norm, window).ratio()
                if score > local_best:
                    local_best = score
                if local_best >= 0.995:
                    break
            if local_best >= 0.995:
                break

        if local_best >= threshold:
            key = (local_best, len(alias_norm))
            if key > best_fuzzy_key:
                best_fuzzy_key = key
                best_fuzzy = {
                    "chemical_name": rec["chemical_name"],
                    "matched_alias": rec["alias"],
                    "match_score": local_best,
                    "match_type": "fuzzy",
                }

    return best_fuzzy


def build_pages_chemical_dataset(
    base_df: pd.DataFrame,
    precursor_workbook: Path,
    output_csv: Path,
) -> None:
    alias_records = load_precursor_alias_records(precursor_workbook)
    token_to_alias_ids, max_alias_token_len = build_alias_index(alias_records)

    scoped = base_df.dropna(subset=["year", "state_abbr"]).copy()
    scoped["year"] = scoped["year"].astype(int)

    matched_rows: list[dict[str, object]] = []
    for row in scoped.itertuples(index=False):
        match = match_chemical_from_goods_description(
            row.goods_description,
            alias_records,
            token_to_alias_ids,
            max_alias_token_len,
        )
        if match is None:
            continue
        matched_rows.append(
            {
                "state_abbr": row.state_abbr,
                "state_name": row.state_name,
                "year": row.year,
                "chemical_name": match["chemical_name"],
                "hs6": row.hs6 if isinstance(row.hs6, str) and row.hs6.strip() else "UNKNOWN",
                "matched_alias": match["matched_alias"],
                "match_type": match["match_type"],
                "match_score": float(match["match_score"]),
                "quantity_kg": row.quantity_kg,
                "value_usd": row.value_usd,
            }
        )

    if matched_rows:
        matched_df = pd.DataFrame(matched_rows)
        grouped = (
            matched_df.groupby(["state_abbr", "state_name", "year", "chemical_name", "hs6"], as_index=False)
            .agg(
                shipment_records=("chemical_name", "count"),
                total_quantity_kg=("quantity_kg", "sum"),
                total_value_usd=("value_usd", "sum"),
                avg_match_score=("match_score", "mean"),
                max_match_score=("match_score", "max"),
            )
            .fillna({"total_quantity_kg": 0.0, "total_value_usd": 0.0})
            .sort_values(["year", "state_abbr", "chemical_name"])
            .reset_index(drop=True)
        )
    else:
        grouped = pd.DataFrame(
            columns=[
                "state_abbr",
                "state_name",
                "year",
                "chemical_name",
                "hs6",
                "shipment_records",
                "total_quantity_kg",
                "total_value_usd",
                "avg_match_score",
                "max_match_score",
            ]
        )

    output_csv.parent.mkdir(parents=True, exist_ok=True)
    grouped.to_csv(output_csv, index=False)

    print(f"Wrote: {output_csv}")
    print(f"Alias entries from precursor list: {len(alias_records):,}")
    print(f"Matched source records: {len(matched_rows):,}")
    print(f"Rows: {len(grouped):,}")
    if len(grouped):
        print(f"Unique chemicals: {grouped['chemical_name'].nunique():,}")
        print(f"Unique HS6 (incl UNKNOWN): {grouped['hs6'].nunique():,}")
        print(f"Year range: {grouped['year'].min()}-{grouped['year'].max()}")


def main() -> None:
    root = Path(__file__).resolve().parent.parent
    input_csv = root / "cnx_transactions_us_sender_or_receiver.csv"
    precursor_workbook = root / "Fentanyl Data" / "Fentanyl_Precursor_List_Combined_with_schedule_date.xlsx"

    base_df = load_base_shipments(input_csv)

    hs_output_csv = root / "docs" / "cnx_shipments_us_state_year_hs6.csv"
    build_pages_hs6_dataset(base_df, hs_output_csv)

    chemical_output_csv = root / "docs" / "cnx_shipments_us_state_year_chemical_matches.csv"
    build_pages_chemical_dataset(base_df, precursor_workbook, chemical_output_csv)


if __name__ == "__main__":
    main()
